# backend > apps > attendance > views.py
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill 
from django.http import HttpResponse
from apps.users.permissions import IsAdmin
import os
import uuid

from django.conf import settings
from django.shortcuts import get_object_or_404
from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import AttendanceSession, AttendanceRecord
from .serializers import (
    AttendanceSessionSerializer,
    AttendanceSessionListSerializer,
)
from .tasks import start_attendance_processing

from .models import AttendanceSession, AttendanceRecord
from apps.classrooms.models import Classroom


ALLOWED_EXTENSIONS = {'.mp4', '.avi', '.mov', '.mkv'}
MAX_UPLOAD_SIZE = getattr(settings, 'MAX_UPLOAD_SIZE', 500 * 1024 * 1024)


class SessionListCreateView(generics.ListCreateAPIView):
    """
    Lista y crea sesiones de asistencia.

    Los maestros solo ven sus propias sesiones; los admins ven todas.
    Al crear, se valida que el aula pertenezca al maestro que solicita.
    """

    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        """Determina el serializador basado en la acción (Lectura vs Escritura)."""
        if self.request.method == 'GET':
            return AttendanceSessionListSerializer
        return AttendanceSessionSerializer

    def get_queryset(self):
        """Optimiza la consulta con select_related y filtra por permisos de usuario."""
        qs = AttendanceSession.objects.select_related(
            'classroom', 'maestro'
        ).prefetch_related('records').filter(classroom__is_active=True)
        if not self.request.user.is_admin:
            qs = qs.filter(maestro=self.request.user)
        return qs.order_by('-created_at')


    
    def perform_create(self, serializer):
        """
        Finaliza la creación de la sesión asignando el maestro y validando permisos.

        Antes de guardar, verifica que si el usuario no es administrador, el aula 
        seleccionada pertenezca obligatoriamente a su cuenta para evitar que un 
        maestro cree sesiones en grupos ajenos.

        Args:
            serializer (AttendanceSessionSerializer): El serializador con los datos 
                ya validados del aula y la fecha.

        Raises:
            PermissionDenied: Si un maestro intenta registrar una sesión en un 
                aula que no tiene asignada.
        """
        classroom = serializer.validated_data['classroom']
        if not self.request.user.is_admin and classroom.maestro != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("No tienes acceso a este grupo.")
        serializer.save(maestro=self.request.user)


class SessionDetailView(generics.RetrieveDestroyAPIView):
    """
    Detalle y eliminación de una sesión específica.

    Attributes:
        serializer_class: AttendanceSessionSerializer para incluir registros.
    """
    serializer_class = AttendanceSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Define la consulta base para obtener las sesiones de asistencia.

        Aplica optimizaciones de base de datos para evitar el problema de consultas 
        N+1 y aplica filtros de seguridad basados en el rol del usuario (Maestro vs Admin).

        Returns:
            QuerySet: Un conjunto de objetos AttendanceSession filtrados y optimizados
                con select_related (claves foráneas) y prefetch_related (relaciones inversas).
        """
        qs = AttendanceSession.objects.select_related(
            'classroom', 'maestro'
        ).prefetch_related('records__student').filter(classroom__is_active=True)
        if not self.request.user.is_admin:
            qs = qs.filter(maestro=self.request.user)
        return qs

    def perform_destroy(self, instance):
        """
        Impide la eliminación si la sesión está en proceso.
        
        Raises:
            ValidationError: Si el estado de la sesión es STATUS_PROCESSING.
        """
        if instance.status == AttendanceSession.STATUS_PROCESSING:
            from rest_framework.exceptions import ValidationError
            raise ValidationError('No se puede eliminar una sesión mientras se está procesando.')
        instance.delete()


class UploadVideoView(APIView):
    """
    Maneja la subida de archivos de video para procesamiento de IA.

    Valida el estado de la sesión, el formato del archivo, el tamaño máximo
    y delega el procesamiento pesado a un hilo de fondo.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, session_id):
        """
        Recibe el video y dispara el pipeline de reconocimiento facial.

        Args:
            request: Objeto de solicitud con el archivo en FILES['video'].
            session_id (int): ID de la sesión a procesar.

        Returns:
            Response: 202 (Accepted) si el video es válido y el proceso inicia.
            Response: 400 (Bad Request) si el video no cumple los requisitos.
        """
        session = self._get_session(request, session_id)

        if session.status == AttendanceSession.STATUS_PROCESSING:
            return Response(
                {'error': 'La sesión está siendo procesada. Espera a que termine.'},
                status=400,
            )
        if session.status == AttendanceSession.STATUS_COMPLETED:
            return Response(
                {'error': 'La sesión ya fue completada. Usa las correcciones manuales si necesitas cambios.'},
                status=400,
            )

        if session.status == AttendanceSession.STATUS_ERROR:
            AttendanceRecord.objects.filter(session=session).delete()
            session.status = AttendanceSession.STATUS_PENDING
            session.error_message = ''
            session.save(update_fields=['status', 'error_message'])

        video_file = request.FILES.get('video')
        if not video_file:
            return Response({'error': 'Se requiere el archivo de video.'}, status=400)

        # Validate extension
        ext = os.path.splitext(video_file.name)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            return Response(
                {'error': f'Formato no permitido. Use: {", ".join(ALLOWED_EXTENSIONS)}'},
                status=400,
            )

        # Validate size
        if video_file.size > MAX_UPLOAD_SIZE:
            return Response({'error': 'El video supera el limite de 500MB.'}, status=400)

        # Save to tmp
        tmp_dir = settings.MEDIA_ROOT / 'tmp'
        tmp_dir.mkdir(parents=True, exist_ok=True)
        filename = f"session_{session.id}_{uuid.uuid4().hex}{ext}"
        video_path = tmp_dir / filename

        with open(video_path, 'wb') as f:
            for chunk in video_file.chunks():
                f.write(chunk)

        # Start background processing
        start_attendance_processing(session.id, str(video_path))

        return Response({
            'message': 'Video recibido. Procesando asistencia en segundo plano.',
            'session_id': session.id,
            'status': AttendanceSession.STATUS_PROCESSING,
        }, status=202)

    def _get_session(self, request, session_id):
        """
        Recupera una sesión específica validando la propiedad y existencia.

        Este método centraliza la lógica de autorización para asegurar que un maestro 
        no pueda subir o manipular videos de una sesión que no le pertenece, incluso 
        si conoce el ID de la sesión.

        Args:
            request (HttpRequest): El objeto de la solicitud que contiene al usuario autenticado.
            session_id (int): El identificador primario de la sesión de asistencia.

        Returns:
            AttendanceSession: La instancia de la sesión si existe y el usuario tiene acceso.

        Raises:
            Http404: Si la sesión no existe o si el usuario intenta acceder a una 
                sesión de otro maestro (evitando fuga de información).
        """
        qs = AttendanceSession.objects.all()
        if not request.user.is_admin:
            qs = qs.filter(maestro=request.user)
        return get_object_or_404(qs, pk=session_id)


class SessionStatusView(APIView):
    """
    Punto de control para consultar el estado actual de procesamiento.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, session_id):
        """
        Retorna el estado de una sesión de asistencia específica.

        Permite verificar si la IA sigue procesando el video, si terminó 
        exitosamente o si ocurrió un error técnico.

        Args:
            request (HttpRequest): Solicitud del usuario autenticado.
            session_id (int): ID de la sesión a consultar.

        Returns:
            Response: JSON con 'status' (pending, processing, completed, error) 
                y el 'error_message' en caso de fallas.
        """
        qs = AttendanceSession.objects.all()
        if not request.user.is_admin:
            qs = qs.filter(maestro=request.user)
        session = get_object_or_404(qs, pk=session_id)
        return Response({
            'session_id': session.id,
            'status': session.status,
            'error_message': session.error_message,
        })


class SessionDeleteView(APIView):
    """
    Maneja la eliminación física de sesiones de asistencia.
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request, session_id):
        """
        Elimina de forma permanente una sesión y sus registros asociados.

        Incluye una validación de integridad para evitar la eliminación de 
        sesiones activas en el pipeline de reconocimiento facial, lo cual 
        podría causar errores de base de datos o archivos huérfanos.

        Args:
            request (HttpRequest): Solicitud del usuario autenticado.
            session_id (int): ID de la sesión a eliminar.

        Returns:
            Response: 204 (No Content) si la eliminación fue exitosa.
            Response: 400 (Bad Request) si la sesión está en estado 'processing'.

        Raises:
            Http404: Si la sesión no existe o no pertenece al maestro solicitante.
        """
        qs = AttendanceSession.objects.all()
        if not request.user.is_admin:
            qs = qs.filter(maestro=request.user)
        session = get_object_or_404(qs, pk=session_id)

        if session.status == AttendanceSession.STATUS_PROCESSING:
            return Response(
                {'error': 'No se puede eliminar una sesión en procesamiento.'},
                status=400,
            )

        session.delete()
        return Response(status=204)


class AttendanceRecordToggleView(APIView):
    """
    Permite la corrección manual de la asistencia.
    Invierte el valor del campo 'is_present' de un registro individual.
    """
    permission_classes = [IsAuthenticated]

    def patch(self, request, record_id):
        """
        Invierte el estado de asistencia de un registro específico (Toggle).

        Permite a los maestros corregir manualmente si un alumno fue marcado 
        erróneamente por la IA. Si estaba como 'Presente' cambia a 'Ausente' 
        y viceversa.

        Args:
            request (HttpRequest): Objeto de la solicitud.
            record_id (int): ID único del AttendanceRecord a modificar.

        Returns:
            Response: JSON con el ID del registro y su nuevo estado 'is_present'.
        """
        qs = AttendanceRecord.objects.select_related('session__classroom')
        if not request.user.is_admin:
            qs = qs.filter(session__maestro=request.user)
        record = get_object_or_404(qs, pk=record_id)

        record.is_present = not record.is_present
        record.save(update_fields=['is_present'])

        return Response({
            'id': record.id,
            'is_present': record.is_present,
        })

class AttendanceExcelReportView(APIView):
    """
    Genera un reporte consolidado en formato .xlsx.

    Crea una matriz donde las filas son alumnos y las columnas son fechas 
    de sesiones completadas, marcando 'P' para presente y 'A' para ausente.

    Query Params:
        classroom_id (int): El ID del grupo a exportar.

    Returns:
        HttpResponse: Archivo binario con el MIME type de Excel.
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def get(self, request):
        """
        Genera y descarga un archivo Excel con el historial de asistencia del grupo.

        Construye dinámicamente una hoja de cálculo con formato profesional. La estructura
        incluye encabezados estilizados, ajuste automático de columnas y una matriz 
        cruzada de estudiantes vs. fechas de sesiones.

        Args:
            request (HttpRequest): Contiene 'classroom_id' en los parámetros de consulta.

        Returns:
            HttpResponse: Un flujo de datos binarios (stream) que el navegador 
                interpreta como un archivo .xlsx descargable.
        """
        from openpyxl.utils import get_column_letter 
        classroom_id = request.query_params.get('classroom_id')
        
        classroom = get_object_or_404(Classroom, pk=classroom_id)
        
        sessions = AttendanceSession.objects.filter(
            classroom=classroom, 
            status='completed'
        ).order_by('date')
        
        students = classroom.students.all().order_by('name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Control de Asistencia"
        ws.merge_cells('A1:B1')
        ws['A1'] = f"GRUPO: {classroom.name}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([]) 

        headers = ['Matrícula', 'Nombre del Alumno']
        for session in sessions:
            headers.append(session.date.strftime('%d/%m/%Y') if session.date else "S/F")
        ws.append(headers)
        
        header_row_idx = ws.max_row
        for cell in ws[header_row_idx]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for student in students:
            row = [student.matricula, student.name]
            for session in sessions:
                record = AttendanceRecord.objects.filter(session=session, student=student).first()
                row.append("P" if (record and record.is_present) else "A")
            ws.append(row)

        for i, col in enumerate(ws.columns, 1):
            column_letter = get_column_letter(i)
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        val_str = str(cell.value)
                        if len(val_str) > max_length:
                            max_length = len(val_str)
                except AttributeError:
                    continue
            
            ws.column_dimensions[column_letter].width = max_length + 3

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        clean_name = str(classroom.name).replace(" ", "_")
        response['Content-Disposition'] = f'attachment; filename="Asistencia_{clean_name}.xlsx"'
        
        wb.save(response)
        return response